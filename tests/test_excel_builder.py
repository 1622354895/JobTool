from job_tracker.excel_builder import build_workbook
from job_tracker.schema import APPLICATION_COLUMNS


def test_workbook_has_required_sheets_tables_and_headers():
    wb = build_workbook()
    assert wb.sheetnames == ["使用说明", "数据看板", "投递记录", "跟进记录", "选项配置"]
    assert "tblApplications" in wb["投递记录"].tables
    assert "tblFollowUps" in wb["跟进记录"].tables
    assert [cell.value for cell in wb["投递记录"][1]] == APPLICATION_COLUMNS


def test_workbook_has_dashboard_visuals_and_validations():
    wb = build_workbook()
    assert len(wb["数据看板"]._charts) == 4
    assert wb["数据看板"]["A7"].value == "近期待办"
    assert wb["数据看板"]["A8"].value == "跟进日期"
    assert "A8:B8" in wb["数据看板"].merged_cells
    assert "H9:L9" in wb["数据看板"].merged_cells
    assert wb["数据看板"]["Q9"].value.startswith("=IFERROR(AGGREGATE")
    assert [chart.anchor for chart in wb["数据看板"]._charts] == ["A20", "I20", "A35", "I35"]
    rule_count = sum(len(group.rules) for group in wb["投递记录"].conditional_formatting)
    assert rule_count >= 4
    assert len(wb["投递记录"].data_validations.dataValidation) >= 5
    assert wb["投递记录"].freeze_panes == "A2"


def test_workbook_has_professional_view_and_print_settings():
    wb = build_workbook()
    applications = wb["投递记录"]
    follow_ups = wb["跟进记录"]
    dashboard = wb["数据看板"]

    assert applications["A1"].font.name == "Microsoft YaHei"
    assert applications.sheet_view.zoomScale == 90
    assert applications.page_setup.orientation == "landscape"
    assert applications.print_title_rows == "$1:$1"
    assert follow_ups.sheet_view.zoomScale == 90
    assert dashboard.sheet_view.zoomScale == 85


def test_workbook_can_be_saved_and_reloaded(tmp_path):
    from openpyxl import load_workbook

    path = tmp_path / "tracker.xlsx"
    build_workbook().save(path)
    loaded = load_workbook(path, data_only=False)
    assert "tblApplications" in loaded["投递记录"].tables
    assert not loaded._external_links
