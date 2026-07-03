from datetime import date

from openpyxl import load_workbook

from job_tracker.excel_builder import build_workbook
from job_tracker.excel_store import ExcelStore
from job_tracker.models import ApplicationDraft
from job_tracker.parser import parse_message
from job_tracker.schema import JOB_TYPES


def make_store(tmp_path):
    path = tmp_path / "tracker.xlsx"
    build_workbook().save(path)
    return ExcelStore(path, backup_dir=tmp_path / "backups")


def test_default_job_types_include_advance_batch():
    assert "提前批" in JOB_TYPES


def test_custom_direction_is_saved_to_workbook_and_available_for_filtering(tmp_path):
    store = make_store(tmp_path)
    options = store.option_groups()
    options["岗位方向"].append("硬件工程")

    store.save_option_groups(options)
    store.append_applications([
        ApplicationDraft(
            company="华为",
            position="硬件工程师",
            direction="硬件工程",
            applied_date=date(2026, 7, 3),
        ),
    ])

    assert "硬件工程" in store.option_groups()["岗位方向"]
    assert store.search({"direction": "硬件工程"})[0]["公司名称"] == "华为"

    workbook = load_workbook(store.path)
    option_sheet = workbook["选项配置"]
    direction_values = [option_sheet.cell(row, 3).value for row in range(2, option_sheet.max_row + 1)]
    workbook.close()
    assert "硬件工程" in direction_values


def test_renaming_option_updates_existing_application_values(tmp_path):
    store = make_store(tmp_path)
    application_id = store.append_applications([
        ApplicationDraft(
            company="字节跳动",
            position="Agent开发实习生",
            direction="Agent开发",
            applied_date=date(2026, 7, 3),
        ),
    ])[0]

    store.rename_option_value("岗位方向", "Agent开发", "智能体开发")

    row = store.get_by_id(application_id)
    assert row["岗位方向"] == "智能体开发"
    options = store.option_groups()
    assert "智能体开发" in options["岗位方向"]
    assert "Agent开发" not in options["岗位方向"]


def test_deleted_option_still_appears_in_filter_choices_when_records_use_it(tmp_path):
    store = make_store(tmp_path)
    store.append_applications([
        ApplicationDraft(
            company="美团",
            position="Java后端实习生",
            direction="Java后端",
            applied_date=date(2026, 7, 3),
        ),
    ])
    options = store.option_groups()
    options["岗位方向"] = [value for value in options["岗位方向"] if value != "Java后端"]
    store.save_option_groups(options)

    assert "Java后端" not in store.option_groups()["岗位方向"]
    assert "Java后端" in store.option_groups(include_record_values=True)["岗位方向"]


def test_parser_can_query_custom_direction_from_runtime_options():
    result = parse_message(
        "查询所有硬件工程方向",
        today=date(2026, 7, 3),
        options={"岗位方向": ["硬件工程"]},
    )

    assert result.query["direction"] == "硬件工程"
