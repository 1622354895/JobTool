from datetime import date

import pytest

from job_tracker.excel_builder import build_workbook
from job_tracker.excel_store import ExcelStore
from job_tracker.models import ApplicationDraft, FollowUpDraft


@pytest.fixture
def store(tmp_path):
    path = tmp_path / "tracker.xlsx"
    build_workbook().save(path)
    return ExcelStore(path, backup_dir=tmp_path / "backups")


def test_append_keeps_existing_rows_and_creates_backup(store):
    store.append_applications([
        ApplicationDraft(company="腾讯", position="大模型实习", applied_date=date(2026, 6, 18)),
    ])
    store.append_applications([
        ApplicationDraft(company="美团", position="Java后端实习", applied_date=date(2026, 6, 19)),
    ])
    rows = store.search({})
    assert [row["公司名称"] for row in rows] == ["腾讯", "美团"]
    assert list(store.backup_dir.glob("tracker_*.xlsx"))

    from openpyxl import load_workbook
    workbook = load_workbook(store.path)
    assert workbook["投递记录"].row_dimensions[3].height == 25
    workbook.close()


def test_search_update_duplicate_and_follow_up(store):
    store.append_applications([
        ApplicationDraft(
            company="字节跳动",
            position="Agent开发实习生",
            direction="Agent开发",
            status="已投递",
            applied_date=date(2026, 6, 18),
        ),
        ApplicationDraft(
            company="美团",
            position="Java后端实习生",
            direction="Java后端",
            status="一面",
            applied_date=date(2026, 6, 17),
        ),
    ])
    rows = store.search({"direction": "Agent开发", "status": "已投递", "keyword": "字节"})
    assert len(rows) == 1
    target_id = rows[0]["投递编号"]
    assert store.find_duplicates("字节跳动", "Agent开发实习生")

    store.update_application(target_id, {"当前状态": "一面"})
    assert store.get_by_id(target_id)["当前状态"] == "一面"

    store.append_follow_up(target_id, FollowUpDraft(content="HR电话沟通", occurred_date=date(2026, 6, 18)))
    assert len(store.follow_ups(target_id)) == 1


def test_append_1000_rows_preserves_history(store):
    drafts = [
        ApplicationDraft(company=f"公司{i}", position="Agent开发实习生", applied_date=date(2026, 6, 18))
        for i in range(1000)
    ]
    store.append_applications(drafts)
    assert len(store.search({})) == 1000


def test_delete_application_removes_application_and_follow_ups(store):
    application_id = store.append_applications([
        ApplicationDraft(company="腾讯", position="Agent实习", applied_date=date(2026, 6, 18)),
    ])[0]
    store.append_follow_up(application_id, FollowUpDraft(content="HR沟通"))
    store.delete_application(application_id)
    assert store.get_by_id(application_id) is None
    assert store.follow_ups(application_id) == []


def test_delete_last_record_preserves_template_row_style(store):
    first_id = store.append_applications([
        ApplicationDraft(company="腾讯", position="Agent实习", applied_date=date(2026, 6, 18)),
    ])[0]
    store.append_follow_up(first_id, FollowUpDraft(content="第一次沟通"))
    store.delete_application(first_id)

    second_id = store.append_applications([
        ApplicationDraft(company="阿里巴巴", position="大模型实习", applied_date=date(2026, 6, 19)),
    ])[0]
    store.append_follow_up(second_id, FollowUpDraft(content="第二次沟通"))

    from openpyxl import load_workbook

    workbook = load_workbook(store.path)
    application_row = workbook["投递记录"][2]
    follow_up_row = workbook["跟进记录"][2]
    assert application_row[0].font.name == "Microsoft YaHei"
    assert application_row[0].font.sz == 10
    assert follow_up_row[0].font.name == "Microsoft YaHei"
    assert follow_up_row[0].font.sz == 10
    assert workbook["投递记录"].row_dimensions[2].height == 25
    assert workbook["跟进记录"].row_dimensions[2].height == 25
    workbook.close()


def test_follow_up_next_date_updates_application_reminder(store):
    application_id = store.append_applications([
        ApplicationDraft(company="字节跳动", position="Agent开发实习", applied_date=date(2026, 6, 18)),
    ])[0]

    store.append_follow_up(
        application_id,
        FollowUpDraft(content="HR沟通", next_action="发送作品集", next_date=date(2026, 6, 22)),
    )

    assert store.get_by_id(application_id)["下次跟进日期"] == date(2026, 6, 22)


def test_invalid_batch_does_not_create_backup(store):
    with pytest.raises(ValueError, match="公司名称和岗位名称不能为空"):
        store.append_applications([
            ApplicationDraft(company="", position="Agent开发实习"),
        ])

    assert list(store.backup_dir.glob("tracker_*.xlsx")) == []
