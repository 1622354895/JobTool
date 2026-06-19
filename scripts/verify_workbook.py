from pathlib import Path
import sys

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT))

from job_tracker.schema import APPLICATION_COLUMNS, FOLLOW_UP_COLUMNS, SHEET_NAMES


def verify(path: Path) -> list[str]:
    errors = []
    if not path.exists():
        return [f"文件不存在：{path}"]
    wb = load_workbook(path, data_only=False, keep_links=True)
    if wb.sheetnames != SHEET_NAMES:
        errors.append(f"工作表顺序错误：{wb.sheetnames}")
    if [cell.value for cell in wb["投递记录"][1]] != APPLICATION_COLUMNS:
        errors.append("投递记录表头错误")
    if [cell.value for cell in wb["跟进记录"][1]] != FOLLOW_UP_COLUMNS:
        errors.append("跟进记录表头错误")
    if "tblApplications" not in wb["投递记录"].tables:
        errors.append("缺少 tblApplications")
    if "tblFollowUps" not in wb["跟进记录"].tables:
        errors.append("缺少 tblFollowUps")
    rule_count = sum(len(group.rules) for group in wb["投递记录"].conditional_formatting)
    if rule_count < 4:
        errors.append("条件格式少于 4 条")
    if len(wb["数据看板"]._charts) != 4:
        errors.append("数据看板图表数量不是 4 张")
    if wb["投递记录"]["A1"].font.name != "Microsoft YaHei":
        errors.append("投递记录未使用中文字体")
    if wb["投递记录"].sheet_view.zoomScale != 90:
        errors.append("投递记录缩放比例错误")
    if wb["投递记录"].page_setup.orientation != "landscape":
        errors.append("投递记录未设置横向打印")
    if wb._external_links:
        errors.append("工作簿包含外部链接")
    formulas = [cell.value for ws in wb.worksheets for row in ws.iter_rows() for cell in row if isinstance(cell.value, str) and cell.value.startswith("=")]
    if not formulas:
        errors.append("工作簿没有动态公式")
    if any(error in formula for formula in formulas for error in ("#REF!", "#DIV/0!", "#NAME?", "#VALUE!")):
        errors.append("工作簿公式包含错误引用")
    wb.close()
    return errors


if __name__ == "__main__":
    target = Path(sys.argv[1]) if len(sys.argv) > 1 else ROOT / "求职投递管理工具.xlsx"
    problems = verify(target)
    if problems:
        print("\n".join(problems))
        raise SystemExit(1)
    print(f"验证通过：{target}")
