import os
import shutil
import tempfile
from copy import copy
from datetime import date, datetime, timedelta
from pathlib import Path

from openpyxl import load_workbook

from .excel_builder import application_formulas, refresh_workbook_options
from .models import ApplicationDraft, FollowUpDraft
from .options import (
    OPTION_CONFIG_VERSION,
    OPTION_GROUP_ORDER,
    OPTION_META_HEADER,
    OPTION_TARGETS,
    clean_option_values,
    merge_option_values,
    merge_with_default_options,
    normalize_option_groups,
)
from .schema import APPLICATION_COLUMNS, FOLLOW_UP_COLUMNS, OPTION_GROUPS


class WorkbookLockedError(RuntimeError):
    pass


class WorkbookFormatError(RuntimeError):
    pass


APPLICATION_FIELD_MAP = {
    "公司名称": "company", "岗位名称": "position", "岗位类型": "job_type", "岗位方向": "direction",
    "工作地点": "location", "投递渠道": "channel", "招聘链接": "url", "投递日期": "applied_date",
    "当前状态": "status", "优先级": "priority", "截止日期": "deadline", "下次跟进日期": "next_follow_up",
    "联系人": "contact", "联系方式": "contact_info", "简历版本": "resume_version", "薪资信息": "salary",
    "标签": "tags", "备注": "notes",
}

CHANGE_FIELD_MAP = {
    "status": "当前状态", "next_follow_up": "下次跟进日期", "priority": "优先级", "notes": "备注",
    "direction": "岗位方向", "channel": "投递渠道", "location": "工作地点",
}


class ExcelStore:
    def __init__(self, path: str | Path, backup_dir: str | Path | None = None):
        self.path = Path(path)
        self.backup_dir = Path(backup_dir or self.path.parent / "backups")

    def _load(self, data_only: bool = False):
        try:
            wb = load_workbook(self.path, data_only=data_only)
        except PermissionError as exc:
            raise WorkbookLockedError("请先关闭 Excel/WPS 中打开的文件") from exc
        except (OSError, KeyError, ValueError) as exc:
            raise WorkbookFormatError("无法读取工作簿") from exc
        required = {"投递记录", "跟进记录", "数据看板", "选项配置"}
        if not required.issubset(wb.sheetnames):
            wb.close()
            raise WorkbookFormatError("所选文件不是本工具生成的格式")
        headers = [cell.value for cell in wb["投递记录"][1]]
        if headers != APPLICATION_COLUMNS:
            wb.close()
            raise WorkbookFormatError("投递记录表头已被修改")
        return wb

    @staticmethod
    def _header_columns(ws) -> dict[str, int]:
        return {str(cell.value): cell.column for cell in ws[1] if cell.value}

    @staticmethod
    def _option_meta_version(ws) -> str:
        headers = ExcelStore._header_columns(ws)
        column = headers.get(OPTION_META_HEADER)
        return str(ws.cell(2, column).value or "") if column else ""

    @staticmethod
    def _read_option_groups(wb) -> dict[str, list[str]]:
        ws = wb["选项配置"]
        headers = ExcelStore._header_columns(ws)
        groups: dict[str, list[str]] = {}
        for group in OPTION_GROUP_ORDER:
            column = headers.get(group)
            values = []
            if column:
                values = [ws.cell(row, column).value for row in range(2, ws.max_row + 1)]
            groups[group] = clean_option_values(values) or list(OPTION_GROUPS[group])
        return normalize_option_groups(groups)

    @staticmethod
    def _merge_record_option_values(wb, groups: dict[str, list[str]]) -> dict[str, list[str]]:
        merged = {group: list(values) for group, values in groups.items()}
        for group, (sheet_name, header) in OPTION_TARGETS.items():
            ws = wb[sheet_name]
            headers = ExcelStore._header_columns(ws)
            column = headers.get(header)
            if not column:
                continue
            record_values = [
                ws.cell(row, column).value
                for row in range(2, ws.max_row + 1)
                if ws.cell(row, 1).value
            ]
            merged[group] = merge_option_values(merged[group], record_values)
        return normalize_option_groups(merged)

    def _refresh_option_views(self, wb) -> None:
        groups = self._read_option_groups(wb)
        dashboard_groups = self._merge_record_option_values(wb, groups)
        refresh_workbook_options(wb, groups, dashboard_groups)

    def ensure_options_ready(self) -> None:
        wb = self._load()
        version = self._option_meta_version(wb["选项配置"])
        groups = self._read_option_groups(wb)
        if version == OPTION_CONFIG_VERSION:
            wb.close()
            return
        self._backup()
        groups = merge_with_default_options(groups)
        dashboard_groups = self._merge_record_option_values(wb, groups)
        refresh_workbook_options(wb, groups, dashboard_groups)
        self._atomic_save(wb)

    def option_groups(self, include_record_values: bool = False) -> dict[str, list[str]]:
        wb = self._load()
        groups = self._read_option_groups(wb)
        if include_record_values:
            groups = self._merge_record_option_values(wb, groups)
        wb.close()
        return groups

    def save_option_groups(self, option_groups: dict[str, list[str]]) -> None:
        wb = self._load()
        self._backup()
        groups = normalize_option_groups(option_groups)
        dashboard_groups = self._merge_record_option_values(wb, groups)
        refresh_workbook_options(wb, groups, dashboard_groups)
        self._atomic_save(wb)

    def rename_option_value(self, group: str, old_value: str, new_value: str) -> None:
        old_value = old_value.strip()
        new_value = new_value.strip()
        if group not in OPTION_TARGETS:
            raise KeyError(f"不支持的选项组：{group}")
        if not old_value or not new_value:
            raise ValueError("选项名称不能为空")
        wb = self._load()
        self._backup()
        groups = self._read_option_groups(wb)
        updated_values = []
        replaced = False
        for value in groups[group]:
            if value == old_value:
                updated_values.append(new_value)
                replaced = True
            else:
                updated_values.append(value)
        if not replaced:
            updated_values.append(new_value)
        groups[group] = clean_option_values(updated_values)

        sheet_name, header = OPTION_TARGETS[group]
        ws = wb[sheet_name]
        headers = self._header_columns(ws)
        column = headers.get(header)
        if column:
            for row in range(2, ws.max_row + 1):
                if ws.cell(row, column).value == old_value:
                    ws.cell(row, column, new_value)
                    if sheet_name == "投递记录" and "最后更新时间" in headers:
                        ws.cell(row, headers["最后更新时间"], datetime.now())
        dashboard_groups = self._merge_record_option_values(wb, groups)
        refresh_workbook_options(wb, groups, dashboard_groups)
        self._atomic_save(wb)

    def _backup(self) -> None:
        self.backup_dir.mkdir(parents=True, exist_ok=True)
        stamp = datetime.now().strftime("%Y%m%d_%H%M%S_%f")
        shutil.copy2(self.path, self.backup_dir / f"{self.path.stem}_{stamp}{self.path.suffix}")
        backups = sorted(self.backup_dir.glob(f"{self.path.stem}_*{self.path.suffix}"), key=lambda item: item.stat().st_mtime, reverse=True)
        for old in backups[30:]:
            old.unlink(missing_ok=True)

    def _atomic_save(self, wb) -> None:
        handle, temp_name = tempfile.mkstemp(prefix=f".{self.path.stem}_", suffix=".xlsx", dir=self.path.parent)
        os.close(handle)
        temp_path = Path(temp_name)
        try:
            wb.save(temp_path)
            wb.close()
            os.replace(temp_path, self.path)
        except PermissionError as exc:
            raise WorkbookLockedError("请先关闭 Excel/WPS 中打开的文件") from exc
        finally:
            temp_path.unlink(missing_ok=True)

    @staticmethod
    def _copy_row_style(ws, source_row: int, target_row: int, max_col: int) -> None:
        ws.row_dimensions[target_row].height = ws.row_dimensions[source_row].height
        for col in range(1, max_col + 1):
            source = ws.cell(source_row, col)
            target = ws.cell(target_row, col)
            if source.has_style:
                target._style = copy(source._style)
            target.number_format = source.number_format
            target.alignment = copy(source.alignment)

    @staticmethod
    def _clear_row_values(ws, row: int, max_col: int) -> None:
        for col in range(1, max_col + 1):
            cell = ws.cell(row, col)
            cell.value = None
            cell.hyperlink = None
            cell.comment = None

    @staticmethod
    def _next_id(ws, prefix: str, id_col: int = 1) -> str:
        today_prefix = f"{prefix}-{date.today():%Y%m%d}-"
        values = [str(ws.cell(row, id_col).value or "") for row in range(2, ws.max_row + 1)]
        suffixes = [int(value.rsplit("-", 1)[1]) for value in values if value.startswith(today_prefix) and value.rsplit("-", 1)[1].isdigit()]
        return f"{today_prefix}{max(suffixes, default=0) + 1:04d}"

    @staticmethod
    def _first_append_row(ws) -> int:
        if ws.max_row == 2 and not ws.cell(2, 1).value and not ws.cell(2, 2).value:
            return 2
        return ws.max_row + 1

    def append_applications(self, drafts: list[ApplicationDraft]) -> list[str]:
        if not drafts:
            return []
        if any(not draft.company.strip() or not draft.position.strip() for draft in drafts):
            raise ValueError("公司名称和岗位名称不能为空")
        wb = self._load()
        self._backup()
        ws = wb["投递记录"]
        table = ws.tables["tblApplications"]
        created_ids: list[str] = []
        for draft in drafts:
            row = self._first_append_row(ws)
            self._copy_row_style(ws, 2, row, len(APPLICATION_COLUMNS))
            application_id = self._next_id(ws, "APP")
            created_ids.append(application_id)
            values = {
                "投递编号": application_id,
                **{header: getattr(draft, field) for header, field in APPLICATION_FIELD_MAP.items()},
                "创建时间": datetime.now(),
                "最后更新时间": datetime.now(),
            }
            formulas = application_formulas(row)
            for col, header in enumerate(APPLICATION_COLUMNS, 1):
                ws.cell(row, col, formulas.get(header, values.get(header)))
            if draft.url:
                link_cell = ws.cell(row, APPLICATION_COLUMNS.index("招聘链接") + 1)
                link_cell.hyperlink = draft.url
                link_cell.style = "Hyperlink"
            table.ref = f"A1:W{row}"
        self._refresh_option_views(wb)
        self._atomic_save(wb)
        return created_ids

    @staticmethod
    def _row_dict(ws, row: int, headers: list[str]) -> dict[str, object]:
        data = {header: ws.cell(row, col).value for col, header in enumerate(headers, 1)}
        status = str(data.get("当前状态") or "")
        follow_date = data.get("下次跟进日期")
        if status in {"Offer", "已拒绝", "已结束"}:
            data["跟进状态"] = "已结束"
        elif not follow_date:
            data["跟进状态"] = "暂无安排"
        else:
            if isinstance(follow_date, datetime):
                follow_date = follow_date.date()
            if isinstance(follow_date, date):
                data["下次跟进日期"] = follow_date
                data["跟进状态"] = "已逾期" if follow_date < date.today() else "今日跟进" if follow_date == date.today() else "未来7天" if follow_date <= date.today() + timedelta(days=7) else "已安排"
        return data

    def search(self, criteria: dict[str, object]) -> list[dict[str, object]]:
        wb = self._load()
        ws = wb["投递记录"]
        rows = [self._row_dict(ws, row, APPLICATION_COLUMNS) for row in range(2, ws.max_row + 1) if ws.cell(row, 1).value]
        wb.close()
        keyword = str(criteria.get("keyword", "")).strip().lower()
        if keyword:
            searchable = ("公司名称", "岗位名称", "岗位方向", "联系人", "标签", "备注")
            rows = [row for row in rows if any(keyword in str(row.get(field, "")).lower() for field in searchable)]
        mappings = {
            "company": "公司名称", "position": "岗位名称", "direction": "岗位方向", "status": "当前状态",
            "channel": "投递渠道", "location": "工作地点", "priority": "优先级",
        }
        for key, header in mappings.items():
            value = criteria.get(key)
            if value:
                rows = [row for row in rows if str(row.get(header, "")).strip().lower() == str(value).strip().lower()]
        if criteria.get("statuses"):
            allowed = set(criteria["statuses"])
            rows = [row for row in rows if row.get("当前状态") in allowed]
        if criteria.get("follow_up_statuses"):
            allowed = set(criteria["follow_up_statuses"])
            rows = [row for row in rows if row.get("跟进状态") in allowed]
        if criteria.get("date_scope") == "this_week":
            start = date.today() - timedelta(days=date.today().weekday())
            rows = [row for row in rows if isinstance(row.get("投递日期"), (date, datetime)) and start <= (row["投递日期"].date() if isinstance(row["投递日期"], datetime) else row["投递日期"]) <= date.today()]
        return rows

    def get_by_id(self, application_id: str) -> dict[str, object] | None:
        rows = self.search({})
        return next((row for row in rows if row["投递编号"] == application_id), None)

    def find_duplicates(self, company: str, position: str) -> list[dict[str, object]]:
        return self.search({"company": company, "position": position})

    def update_application(self, application_id: str, changes: dict[str, object]) -> None:
        wb = self._load()
        ws = wb["投递记录"]
        target_row = next((row for row in range(2, ws.max_row + 1) if ws.cell(row, 1).value == application_id), None)
        if not target_row:
            wb.close()
            raise KeyError(f"找不到投递记录：{application_id}")
        self._backup()
        for key, value in changes.items():
            header = CHANGE_FIELD_MAP.get(key, key)
            if header not in APPLICATION_COLUMNS or header in {"投递编号", "创建时间"}:
                continue
            ws.cell(target_row, APPLICATION_COLUMNS.index(header) + 1, value)
        ws.cell(target_row, APPLICATION_COLUMNS.index("最后更新时间") + 1, datetime.now())
        self._refresh_option_views(wb)
        self._atomic_save(wb)

    def append_follow_up(self, application_id: str, follow_up: FollowUpDraft) -> str:
        wb = self._load()
        application_ws = wb["投递记录"]
        application_row = next(
            (row for row in range(2, application_ws.max_row + 1) if application_ws.cell(row, 1).value == application_id),
            None,
        )
        if not application_row:
            wb.close()
            raise KeyError(f"找不到投递记录：{application_id}")
        self._backup()
        ws = wb["跟进记录"]
        row = self._first_append_row(ws)
        self._copy_row_style(ws, 2, row, len(FOLLOW_UP_COLUMNS))
        record_id = self._next_id(ws, "LOG")
        values = {
            "记录编号": record_id, "投递编号": application_id,
            "公司名称": application_ws.cell(application_row, APPLICATION_COLUMNS.index("公司名称") + 1).value,
            "岗位名称": application_ws.cell(application_row, APPLICATION_COLUMNS.index("岗位名称") + 1).value,
            "记录类型": follow_up.record_type,
            "发生日期": follow_up.occurred_date or date.today(), "过程内容": follow_up.content,
            "面试问题": follow_up.interview_questions, "自我复盘": follow_up.self_review,
            "对方反馈": follow_up.feedback, "下一步行动": follow_up.next_action,
            "下一步日期": follow_up.next_date,
        }
        for col, header in enumerate(FOLLOW_UP_COLUMNS, 1):
            ws.cell(row, col, values.get(header))
        ws.tables["tblFollowUps"].ref = f"A1:L{row}"
        if follow_up.next_date:
            application_ws.cell(
                application_row,
                APPLICATION_COLUMNS.index("下次跟进日期") + 1,
                follow_up.next_date,
            )
            application_ws.cell(
                application_row,
                APPLICATION_COLUMNS.index("最后更新时间") + 1,
                datetime.now(),
            )
        self._atomic_save(wb)
        return record_id

    def delete_application(self, application_id: str) -> None:
        wb = self._load()
        app_ws = wb["投递记录"]
        target_row = next((row for row in range(2, app_ws.max_row + 1) if app_ws.cell(row, 1).value == application_id), None)
        if not target_row:
            wb.close()
            raise KeyError(f"找不到投递记录：{application_id}")
        self._backup()
        application_rows = [row for row in range(2, app_ws.max_row + 1) if app_ws.cell(row, 1).value]
        if len(application_rows) == 1:
            self._clear_row_values(app_ws, target_row, len(APPLICATION_COLUMNS))
            for header, formula in application_formulas(target_row).items():
                app_ws.cell(target_row, APPLICATION_COLUMNS.index(header) + 1, formula)
        else:
            app_ws.delete_rows(target_row)
        app_last = max(2, app_ws.max_row)
        app_ws.tables["tblApplications"].ref = f"A1:W{app_last}"
        follow_ws = wb["跟进记录"]
        for row in range(follow_ws.max_row, 1, -1):
            if follow_ws.cell(row, 2).value == application_id:
                follow_rows = [item for item in range(2, follow_ws.max_row + 1) if follow_ws.cell(item, 1).value]
                if len(follow_rows) == 1:
                    self._clear_row_values(follow_ws, row, len(FOLLOW_UP_COLUMNS))
                else:
                    follow_ws.delete_rows(row)
        follow_last = max(2, follow_ws.max_row)
        follow_ws.tables["tblFollowUps"].ref = f"A1:L{follow_last}"
        self._refresh_option_views(wb)
        self._atomic_save(wb)

    def follow_ups(self, application_id: str | None = None) -> list[dict[str, object]]:
        wb = self._load()
        ws = wb["跟进记录"]
        rows = [
            {header: ws.cell(row, col).value for col, header in enumerate(FOLLOW_UP_COLUMNS, 1)}
            for row in range(2, ws.max_row + 1)
            if ws.cell(row, 1).value
        ]
        wb.close()
        return [row for row in rows if row["投递编号"] == application_id] if application_id else rows
