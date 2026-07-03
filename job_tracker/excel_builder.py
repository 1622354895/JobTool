from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.page import PageMargins
from openpyxl.worksheet.table import Table, TableStyleInfo

from .options import (
    OPTION_CONFIG_VERSION,
    OPTION_GROUP_ORDER,
    OPTION_META_HEADER,
    normalize_option_groups,
)
from .schema import APPLICATION_COLUMNS, FOLLOW_UP_COLUMNS, OPTION_GROUPS, SHEET_NAMES


NAVY = "17324D"
BLUE = "1677D2"
LIGHT_BLUE = "EAF4FC"
LIGHT_GRAY = "F4F6F8"
GREEN = "2E8B57"
LIGHT_GREEN = "E8F5EE"
AMBER = "D78A00"
LIGHT_AMBER = "FFF3D6"
RED = "C83F49"
LIGHT_RED = "FDEBEC"
WHITE = "FFFFFF"
TEXT = "203040"
GRID = "D9E1E8"
FONT_NAME = "Microsoft YaHei"


def _table_style() -> TableStyleInfo:
    return TableStyleInfo(name="TableStyleMedium2", showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=False)


def _style_table_sheet(ws, headers: list[str], widths: dict[str, int]) -> None:
    ws.freeze_panes = "A2"
    ws.sheet_view.showGridLines = False
    ws.sheet_view.zoomScale = 90
    ws.sheet_view.zoomScaleNormal = 90
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.print_title_rows = "$1:$1"
    ws.page_margins = PageMargins(left=0.25, right=0.25, top=0.5, bottom=0.5, header=0.2, footer=0.2)
    ws.auto_filter.ref = f"A1:{ws.cell(1, len(headers)).column_letter}2"
    for cell in ws[1]:
        cell.font = Font(name=FONT_NAME, size=10, bold=True, color=WHITE)
        cell.fill = PatternFill("solid", fgColor=NAVY)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 30
    for index, header in enumerate(headers, 1):
        letter = ws.cell(1, index).column_letter
        ws.column_dimensions[letter].width = widths.get(header, 14)
    for col in range(1, len(headers) + 1):
        cell = ws.cell(2, col)
        cell.font = Font(name=FONT_NAME, size=10, color=TEXT)
        cell.alignment = Alignment(vertical="center", wrap_text=False)
        cell.border = Border(bottom=Side(style="hair", color=GRID))
    ws.row_dimensions[2].height = 25
    for name in ("投递日期", "截止日期", "下次跟进日期", "创建时间", "最后更新时间", "发生日期", "下一步日期"):
        if name in headers:
            col = headers.index(name) + 1
            ws.cell(2, col).number_format = "yyyy-mm-dd"


def _option_column(ws, group: str) -> int | None:
    for cell in ws[1]:
        if cell.value == group:
            return cell.column
    return None


def _option_range(wb, group: str, option_groups: dict[str, list[str]]) -> str:
    ws = wb["选项配置"]
    column = _option_column(ws, group)
    if column is None:
        raise KeyError(f"缺少选项组：{group}")
    letter = ws.cell(1, column).column_letter
    end_row = max(2, len(option_groups[group]) + 1)
    return f"'选项配置'!${letter}$2:${letter}${end_row}"


def _clear_validations(ws) -> None:
    ws.data_validations.dataValidation = []


def apply_option_validations(wb, option_groups: dict[str, list[str]] | None = None) -> None:
    option_groups = normalize_option_groups(option_groups or OPTION_GROUPS)
    ws = wb["投递记录"]
    _clear_validations(ws)
    mappings = {
        "岗位类型": "岗位类型",
        "岗位方向": "岗位方向",
        "投递渠道": "投递渠道",
        "当前状态": "状态",
        "优先级": "优先级",
    }
    headers = {cell.value: cell.column_letter for cell in ws[1]}
    for header, group in mappings.items():
        validation = DataValidation(type="list", formula1=_option_range(wb, group, option_groups), allow_blank=True)
        validation.error = "请选择下拉列表中的值"
        validation.errorTitle = "无效选项"
        validation.prompt = f"请选择{header}"
        validation.promptTitle = header
        ws.add_data_validation(validation)
        validation.add(f"{headers[header]}2:{headers[header]}5000")
    date_validation = DataValidation(type="date", operator="between", formula1="DATE(2000,1,1)", formula2="DATE(2100,12,31)", allow_blank=True)
    date_validation.error = "请输入有效日期"
    ws.add_data_validation(date_validation)
    for header in ("投递日期", "截止日期", "下次跟进日期"):
        letter = headers[header]
        date_validation.add(f"{letter}2:{letter}5000")

    follow_ws = wb["跟进记录"]
    _clear_validations(follow_ws)
    log_validation = DataValidation(type="list", formula1=_option_range(wb, "记录类型", option_groups), allow_blank=True)
    log_validation.error = "请选择下拉列表中的值"
    log_validation.errorTitle = "无效选项"
    follow_ws.add_data_validation(log_validation)
    log_validation.add("E2:E5000")


def _add_conditional_formatting(ws) -> None:
    ws.conditional_formatting.add("A2:W5000", FormulaRule(formula=['$N2="已逾期"'], fill=PatternFill("solid", fgColor=LIGHT_RED)))
    ws.conditional_formatting.add("A2:W5000", FormulaRule(formula=['$N2="今日跟进"'], fill=PatternFill("solid", fgColor=LIGHT_AMBER)))
    ws.conditional_formatting.add("A2:W5000", FormulaRule(formula=['$J2="Offer"'], fill=PatternFill("solid", fgColor=LIGHT_GREEN)))
    ws.conditional_formatting.add("A2:W5000", FormulaRule(formula=['$K2="高"'], font=Font(name=FONT_NAME, bold=True, color=AMBER)))
    ws.conditional_formatting.add("A2:W5000", FormulaRule(formula=['$U2="疑似重复"'], font=Font(name=FONT_NAME, bold=True, color=RED)))


def application_formulas(row: int) -> dict[str, str]:
    return {
        "跟进状态": f'=IF(OR(J{row}="Offer",J{row}="已拒绝",J{row}="已结束"),"已结束",IF(M{row}="","暂无安排",IF(M{row}<TODAY(),"已逾期",IF(M{row}=TODAY(),"今日跟进",IF(M{row}<=TODAY()+7,"未来7天","已安排")))))',
        "重复检查": f'=IF(OR(B{row}="",C{row}=""),"",IF(COUNTIFS(tblApplications[公司名称],B{row},tblApplications[岗位名称],C{row})>1,"疑似重复",""))',
    }


def _build_instructions(ws) -> None:
    ws.sheet_view.showGridLines = False
    ws.sheet_view.zoomScale = 95
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1
    ws.column_dimensions["A"].width = 4
    ws.column_dimensions["B"].width = 11
    for column in ("C", "D", "E", "F", "G", "H"):
        ws.column_dimensions[column].width = 16
    ws.merge_cells("B2:H3")
    ws["B2"] = "求职投递记录助手"
    ws["B2"].font = Font(name=FONT_NAME, size=24, bold=True, color=WHITE)
    ws["B2"].fill = PatternFill("solid", fgColor=NAVY)
    ws["B2"].alignment = Alignment(horizontal="left", vertical="center")
    for row in ws["B2:H3"]:
        for cell in row:
            cell.fill = PatternFill("solid", fgColor=NAVY)
    ws.row_dimensions[2].height = 30
    ws.row_dimensions[3].height = 30
    ws.merge_cells("B5:H5")
    ws["B5"] = "快速开始"
    ws["B5"].font = Font(name=FONT_NAME, size=12, bold=True, color=BLUE)
    ws["B5"].fill = PatternFill("solid", fgColor=LIGHT_BLUE)
    ws["B5"].alignment = Alignment(vertical="center")
    items = [
        ("01", "打开桌面程序，单条岗位使用“快速表单”，批量信息使用“智能消息”。"),
        ("02", "解析或填写完成后先核对预览，再点击“确认写入 Excel”。"),
        ("03", "在“投递记录”中搜索、筛选和更新状态，在“跟进中心”处理待办。"),
        ("04", "写入前请关闭 Excel/WPS 中打开的当前文件，程序会自动创建备份。"),
        ("05", "红色代表逾期，黄色代表今日跟进，绿色代表 Offer。"),
        ("06", "不要修改工作表名称、表头和公式列，以免影响程序读取。"),
    ]
    for row, (number, text) in enumerate(items, 7):
        ws.cell(row, 2, number).font = Font(name=FONT_NAME, size=10, bold=True, color=WHITE)
        ws.cell(row, 2).fill = PatternFill("solid", fgColor=BLUE)
        ws.cell(row, 2).alignment = Alignment(horizontal="center", vertical="center")
        ws.merge_cells(start_row=row, start_column=3, end_row=row, end_column=8)
        ws.cell(row, 3, text).font = Font(name=FONT_NAME, size=10, color=TEXT)
        ws.cell(row, 3).alignment = Alignment(vertical="center", wrap_text=True)
        ws.cell(row, 3).fill = PatternFill("solid", fgColor="F8FAFC")
        ws.row_dimensions[row].height = 32
    ws.merge_cells("B15:H15")
    ws["B15"] = "推荐输入：公司：字节跳动；岗位：Agent开发实习生；日期：今天；状态：已投递；渠道：Boss直聘"
    ws["B15"].font = Font(name=FONT_NAME, size=10, italic=True, color="5B6B79")
    ws["B15"].fill = PatternFill("solid", fgColor=LIGHT_AMBER)
    ws["B15"].alignment = Alignment(vertical="center", wrap_text=True)
    ws.row_dimensions[15].height = 38
    ws["B17"] = "文件生成时间"
    ws["B17"].font = Font(name=FONT_NAME, size=9, color="667786")
    ws["C17"] = datetime.now()
    ws["C17"].font = Font(name=FONT_NAME, size=9, color="667786")
    ws["C17"].number_format = "yyyy-mm-dd hh:mm"


def write_options_sheet(ws, option_groups: dict[str, list[str]] | None = None) -> None:
    option_groups = normalize_option_groups(option_groups or OPTION_GROUPS)
    if ws.max_row:
        ws.delete_rows(1, ws.max_row)
    ws.sheet_view.showGridLines = False
    ws.sheet_view.zoomScale = 95
    for col, name in enumerate(OPTION_GROUP_ORDER, 1):
        values = option_groups[name]
        ws.cell(1, col, name)
        ws.cell(1, col).font = Font(name=FONT_NAME, bold=True, color=WHITE)
        ws.cell(1, col).fill = PatternFill("solid", fgColor=NAVY)
        ws.column_dimensions[ws.cell(1, col).column_letter].width = 18
        for row, value in enumerate(values, 2):
            ws.cell(row, col, value)
            ws.cell(row, col).font = Font(name=FONT_NAME, color=TEXT)
            ws.cell(row, col).fill = PatternFill("solid", fgColor=WHITE if row % 2 == 0 else LIGHT_GRAY)
    meta_col = len(OPTION_GROUP_ORDER) + 2
    ws.cell(1, meta_col, OPTION_META_HEADER)
    ws.cell(2, meta_col, OPTION_CONFIG_VERSION)
    ws.column_dimensions[ws.cell(1, meta_col).column_letter].hidden = True


def _build_config(ws, option_groups: dict[str, list[str]] | None = None) -> None:
    write_options_sheet(ws, option_groups)


def _metric_card(ws, label_cell: str, value_cell: str, label: str, formula: str, number_format: str = "0") -> None:
    ws[label_cell] = label
    ws[label_cell].font = Font(name=FONT_NAME, size=10, bold=True, color="5B6B79")
    ws[value_cell] = formula
    ws[value_cell].font = Font(name=FONT_NAME, size=20, bold=True, color=NAVY)
    ws[value_cell].number_format = number_format
    for cell in (ws[label_cell], ws[value_cell]):
        cell.fill = PatternFill("solid", fgColor=WHITE)
        cell.border = Border(bottom=Side(style="thin", color=GRID))
        cell.alignment = Alignment(horizontal="center", vertical="center")


def _build_dashboard(ws, option_groups: dict[str, list[str]] | None = None) -> None:
    option_groups = normalize_option_groups(option_groups or OPTION_GROUPS)
    ws.sheet_view.showGridLines = False
    ws.sheet_view.zoomScale = 85
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1
    for col in range(1, 18):
        ws.column_dimensions[ws.cell(1, col).column_letter].width = 12
    ws.merge_cells("A1:P1")
    ws["A1"] = "求职进展总览"
    ws["A1"].font = Font(name=FONT_NAME, size=22, bold=True, color=WHITE)
    ws["A1"].alignment = Alignment(vertical="center")
    for row in ws["A1:P1"]:
        for cell in row:
            cell.fill = PatternFill("solid", fgColor=NAVY)
    ws.row_dimensions[1].height = 42
    metrics = [
        ("A3", "A4", "累计投递", '=COUNTIF(tblApplications[公司名称],"<>")', "0"),
        ("C3", "C4", "本周投递", '=COUNTIFS(tblApplications[投递日期],">="&TODAY()-WEEKDAY(TODAY(),2)+1,tblApplications[投递日期],"<="&TODAY())', "0"),
        ("E3", "E4", "本月投递", '=COUNTIFS(tblApplications[投递日期],">="&EOMONTH(TODAY(),-1)+1,tblApplications[投递日期],"<="&TODAY())', "0"),
        ("G3", "G4", "待跟进", '=COUNTIF(tblApplications[跟进状态],"已逾期")+COUNTIF(tblApplications[跟进状态],"今日跟进")', "0"),
        ("I3", "I4", "面试中", '=SUM(COUNTIF(tblApplications[当前状态],{"一面","二面","终面"}))', "0"),
        ("K3", "K4", "Offer", '=COUNTIF(tblApplications[当前状态],"Offer")', "0"),
        ("M3", "M4", "回复率", '=IFERROR(SUM(COUNTIF(tblApplications[当前状态],{"初筛沟通","笔试中","一面","二面","终面","等待结果","Offer"}))/COUNTIF(tblApplications[公司名称],"<>"),0)', "0.0%"),
        ("O3", "O4", "面试转化", '=IFERROR(SUM(COUNTIF(tblApplications[当前状态],{"一面","二面","终面","等待结果","Offer"}))/COUNTIF(tblApplications[公司名称],"<>"),0)', "0.0%"),
    ]
    for args in metrics:
        _metric_card(ws, *args)
    ws.merge_cells("A7:P7")
    ws["A7"] = "近期待办"
    ws["A7"].font = Font(name=FONT_NAME, size=12, bold=True, color=NAVY)
    ws["A7"].fill = PatternFill("solid", fgColor=LIGHT_BLUE)
    ws["A7"].alignment = Alignment(vertical="center")
    ws.row_dimensions[7].height = 28
    todo_layout = [
        (1, 2, "跟进日期", "下次跟进日期"),
        (3, 4, "跟进状态", "跟进状态"),
        (5, 7, "公司名称", "公司名称"),
        (8, 12, "岗位名称", "岗位名称"),
        (13, 14, "优先级", "优先级"),
        (15, 16, "当前状态", "当前状态"),
    ]
    for start_col, end_col, header, _source_header in todo_layout:
        ws.merge_cells(start_row=8, start_column=start_col, end_row=8, end_column=end_col)
        cell = ws.cell(8, start_col, header)
        cell.font = Font(name=FONT_NAME, size=9, bold=True, color=WHITE)
        cell.fill = PatternFill("solid", fgColor=NAVY)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        for col in range(start_col, end_col + 1):
            ws.cell(8, col).fill = PatternFill("solid", fgColor=NAVY)
    ws.row_dimensions[8].height = 24
    for row in range(9, 19):
        rank = row - 8
        ws.cell(row, 17, (
            '=IFERROR(AGGREGATE(15,6,'
            '(ROW(tblApplications[下次跟进日期])-ROW(INDEX(tblApplications[下次跟进日期],1,1))+1)/'
            '((tblApplications[下次跟进日期]<>"")*(tblApplications[当前状态]<>"Offer")*'
            '(tblApplications[当前状态]<>"已拒绝")*(tblApplications[当前状态]<>"已结束")),'
            f'{rank}),"")'
        ))
        for start_col, end_col, _header, source_header in todo_layout:
            ws.merge_cells(start_row=row, start_column=start_col, end_row=row, end_column=end_col)
            ws.cell(row, start_col, f'=IF($Q{row}="","",INDEX(tblApplications[{source_header}],$Q{row}))')
            ws.cell(row, start_col).font = Font(name=FONT_NAME, size=9, color=TEXT)
            ws.cell(row, start_col).alignment = Alignment(
                horizontal="center" if source_header in {"下次跟进日期", "跟进状态", "优先级", "当前状态"} else "left",
                vertical="center",
                wrap_text=source_header == "岗位名称",
            )
            for col in range(start_col, end_col + 1):
                ws.cell(row, col).fill = PatternFill("solid", fgColor=WHITE if row % 2 else LIGHT_GRAY)
                ws.cell(row, col).border = Border(bottom=Side(style="hair", color=GRID))
        ws.cell(row, 1).number_format = "yyyy-mm-dd"
        ws.row_dimensions[row].height = 24
    ws.conditional_formatting.add("A9:P18", FormulaRule(formula=['$C9="已逾期"'], fill=PatternFill("solid", fgColor=LIGHT_RED)))
    ws.conditional_formatting.add("A9:P18", FormulaRule(formula=['$C9="今日跟进"'], fill=PatternFill("solid", fgColor=LIGHT_AMBER)))
    statuses = option_groups["状态"]
    directions = option_groups["岗位方向"]
    channels = option_groups["投递渠道"]
    ws["R1"], ws["S1"] = "状态", "数量"
    for row, value in enumerate(statuses, 2):
        ws.cell(row, 18, value)
        ws.cell(row, 19, f'=COUNTIF(tblApplications[当前状态],R{row})')
    ws["U1"], ws["V1"] = "方向", "数量"
    for row, value in enumerate(directions, 2):
        ws.cell(row, 21, value)
        ws.cell(row, 22, f'=COUNTIF(tblApplications[岗位方向],U{row})')
    ws["X1"], ws["Y1"] = "渠道", "数量"
    for row, value in enumerate(channels, 2):
        ws.cell(row, 24, value)
        ws.cell(row, 25, f'=COUNTIF(tblApplications[投递渠道],X{row})')
    ws["AA1"], ws["AB1"] = "日期", "数量"
    for row in range(2, 32):
        ws.cell(row, 27, f'=TEXT(TODAY()-{32-row},"m-d")')
        ws.cell(row, 28, f'=COUNTIF(tblApplications[投递日期],AA{row})')
    charts = []
    status_chart = BarChart()
    status_chart.type = "bar"
    status_chart.title = "投递状态分布"
    status_max_row = max(2, len(statuses) + 1)
    status_chart.add_data(Reference(ws, min_col=19, min_row=1, max_row=status_max_row), titles_from_data=True)
    status_chart.set_categories(Reference(ws, min_col=18, min_row=2, max_row=status_max_row))
    status_chart.height, status_chart.width = 7, 10
    status_chart.series[0].graphicalProperties.solidFill = BLUE
    charts.append((status_chart, "A20"))
    direction_chart = BarChart()
    direction_chart.type = "bar"
    direction_chart.title = "岗位方向分布"
    direction_max_row = max(2, len(directions) + 1)
    direction_chart.add_data(Reference(ws, min_col=22, min_row=1, max_row=direction_max_row), titles_from_data=True)
    direction_chart.set_categories(Reference(ws, min_col=21, min_row=2, max_row=direction_max_row))
    direction_chart.height, direction_chart.width = 7, 10
    direction_chart.series[0].graphicalProperties.solidFill = GREEN
    charts.append((direction_chart, "I20"))
    channel_chart = BarChart()
    channel_chart.type = "bar"
    channel_chart.title = "投递渠道分布"
    channel_max_row = max(2, len(channels) + 1)
    channel_chart.add_data(Reference(ws, min_col=25, min_row=1, max_row=channel_max_row), titles_from_data=True)
    channel_chart.set_categories(Reference(ws, min_col=24, min_row=2, max_row=channel_max_row))
    channel_chart.height, channel_chart.width = 7, 10
    channel_chart.series[0].graphicalProperties.solidFill = AMBER
    charts.append((channel_chart, "A35"))
    trend_chart = LineChart()
    trend_chart.title = "最近30天投递趋势"
    trend_chart.add_data(Reference(ws, min_col=28, min_row=1, max_row=31), titles_from_data=True)
    trend_chart.set_categories(Reference(ws, min_col=27, min_row=2, max_row=31))
    trend_chart.height, trend_chart.width = 7, 10
    trend_chart.x_axis.tickLblSkip = 5
    trend_chart.series[0].graphicalProperties.line.solidFill = BLUE
    charts.append((trend_chart, "I35"))
    for chart, anchor in charts:
        chart.legend = None
        chart.style = 10
        ws.add_chart(chart, anchor)
    for col in ("Q", "R", "S", "U", "V", "X", "Y", "AA", "AB"):
        ws.column_dimensions[col].hidden = True


def rebuild_dashboard_sheet(wb, option_groups: dict[str, list[str]] | None = None) -> None:
    index = wb.sheetnames.index("数据看板")
    old = wb["数据看板"]
    wb.remove(old)
    ws = wb.create_sheet("数据看板", index)
    _build_dashboard(ws, option_groups)


def refresh_workbook_options(
    wb,
    option_groups: dict[str, list[str]] | None = None,
    dashboard_groups: dict[str, list[str]] | None = None,
) -> dict[str, list[str]]:
    normalized = normalize_option_groups(option_groups or OPTION_GROUPS)
    write_options_sheet(wb["选项配置"], normalized)
    apply_option_validations(wb, normalized)
    rebuild_dashboard_sheet(wb, dashboard_groups or normalized)
    return normalized


def build_workbook(option_groups: dict[str, list[str]] | None = None) -> Workbook:
    option_groups = normalize_option_groups(option_groups or OPTION_GROUPS)
    wb = Workbook()
    wb.remove(wb.active)
    for name in SHEET_NAMES:
        wb.create_sheet(name)
    _build_instructions(wb["使用说明"])
    _build_config(wb["选项配置"], option_groups)
    application_ws = wb["投递记录"]
    application_ws.append(APPLICATION_COLUMNS)
    application_ws.append([None] * len(APPLICATION_COLUMNS))
    formulas = application_formulas(2)
    for header, formula in formulas.items():
        application_ws.cell(2, APPLICATION_COLUMNS.index(header) + 1, formula)
    app_table = Table(displayName="tblApplications", ref=f"A1:W2")
    app_table.tableStyleInfo = _table_style()
    application_ws.add_table(app_table)
    widths = {
        "投递编号": 19, "公司名称": 18, "岗位名称": 26, "岗位类型": 13, "岗位方向": 14,
        "工作地点": 12, "投递渠道": 14, "招聘链接": 24, "投递日期": 13, "当前状态": 13,
        "优先级": 10, "截止日期": 13, "下次跟进日期": 15, "跟进状态": 13, "联系人": 13,
        "联系方式": 18, "简历版本": 15, "薪资信息": 16, "标签": 18, "备注": 30,
        "重复检查": 13, "创建时间": 18, "最后更新时间": 18,
    }
    _style_table_sheet(application_ws, APPLICATION_COLUMNS, widths)
    _add_conditional_formatting(application_ws)
    follow_ws = wb["跟进记录"]
    follow_ws.append(FOLLOW_UP_COLUMNS)
    follow_ws.append([None] * len(FOLLOW_UP_COLUMNS))
    follow_table = Table(displayName="tblFollowUps", ref=f"A1:L2")
    follow_table.tableStyleInfo = _table_style()
    follow_ws.add_table(follow_table)
    _style_table_sheet(follow_ws, FOLLOW_UP_COLUMNS, {"过程内容": 34, "面试问题": 30, "自我复盘": 30, "对方反馈": 25, "下一步行动": 25})
    apply_option_validations(wb, option_groups)
    _build_dashboard(wb["数据看板"], option_groups)
    wb["使用说明"].sheet_view.tabSelected = True
    wb.active = 0
    try:
        wb.calculation.fullCalcOnLoad = True
        wb.calculation.forceFullCalc = True
        wb.calculation.calcMode = "auto"
    except AttributeError:
        pass
    return wb


def save_template(path: str | Path) -> Path:
    output = Path(path)
    output.parent.mkdir(parents=True, exist_ok=True)
    build_workbook().save(output)
    return output
